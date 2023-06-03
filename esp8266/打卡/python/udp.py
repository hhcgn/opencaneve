import json
import os
from socket import *
import openpyxl
import xlrd
import xlwt
import datetime as dat
import subprocess
from openpyxl import load_workbook
from xlrd import open_workbook
from datetime import datetime
# from xlutils.copy import copy
import time


# date_day = str(datetime.datetime.now().year) + "-" + str(datetime.datetime.now().month) + "-" + str(
#     datetime.datetime.now().day)
# savepath = os.path.abspath(os.path.dirname(os.getcwd())) + "/" + date_day + '.xls'
# book = xlwt.Workbook(encoding='utf-8', style_compression=0)
# sheet = book.add_sheet('打卡', cell_overwrite_ok=False)
# col = ('S', 'I', '签到时间', '签退时间', 'ER', 'R','DATE')
# line_i=0
# same=0
# addr_assist=''
# for i in range(0, 7):
#     sheet.write(0, i, col[i])
# book.save(savepath)

read_member=os.path.abspath("D:/RM_Files" + '/member.xlsx')
member=openpyxl.load_workbook(read_member)
member_sheet=member["Sheet1"]
# rows 按照行获取表单中所有的格子，每一行的数据放到一个元祖中
res = list(member_sheet.rows)
# 获取excel表格中的第一行的数据，作为字典的key==》生成一个list列表
title = [i.value for i in res[0]]
# 作为每个字典的容器
cases1 = []
cases2 = []
# 遍历第一行意外的所有行
for item in res[1:]:
    # 获取每行的数据
    dataline = [i.value for i in item]
    # 把遍历的每行数据与第一行title数据打包成字典
    cases1.append(dataline[1])
    cases2.append(dataline[0])
dicline = dict(zip(cases1, cases2))
#print(dicline)
#print(dicline.get("xx"))



def send_msg(udp_socket):
    dest_ip = input("请输入目标IP：")
    dest_port = int(input("请输入端口号："))
    data = input("请输入要发送的内容：")
    udp_socket.sendto(data.encode('GBK'), (dest_ip, dest_port))


def recv_msg(udp_socket):
    print("start")
    while True:
        recv_data = udp_socket.recvfrom(1024)
        original_data = recv_data[0].decode('GBK')
        # print(recv_data)
        

        count_n = 0
        record_line = 1;
        global same, data_dict
        global line_i
        global addr_assist,sendback
        sendback="no"
        date_day = str(dat.datetime.now().year) + "-" + str(dat.datetime.now().month) + "-" + str(
            dat.datetime.now().day)
        savepath = os.path.abspath("D:\RM_Files" +"/"+ date_day + '.xlsx')
        try:
            data_dict = json.loads(original_data)
        except:
            print("格式不正确")
        if (os.path.exists(savepath)):
            wb = openpyxl.load_workbook(savepath)
            wa = wb.active

        else:
            print("none")
            wb = openpyxl.Workbook()  # 创建一个excel文件
            wa = wb.active  # 获得一个的工作表
            wa.title = "打卡"
            wa.append(['S', 'ID', '签到', '签退', 'ER', 'NAME', 'TIME'])

        current_time = str(dat.datetime.now())[-15:-7]

        try:
            for i in range(1, wa.max_row + 1)[::-1]:
                if (wb["打卡"].cell(row=i, column=2).value) == data_dict['I']:
                    record_line = i
                    break

            if wb["打卡"].cell(row=record_line, column=1).value == "ok":
                sendback = "ok"
                print("签到1")
                wa.append(
                ["no", data_dict['I'], current_time, data_dict['W'], data_dict['ER'], dicline.get(data_dict['I']),
                 0])
                print(dicline.get(data_dict['I']))
            elif wb["打卡"].cell(row=record_line, column=1).value == "no":
                print("签退1")
                sendback = "end"  # 反馈给打卡终端数据
                time_1 = str(wb["打卡"].cell(row=record_line, column=3).value)
                time_2 = str(current_time)
                time_1_struct = datetime.strptime(time_1, "%H:%M:%S")
                time_2_struct = datetime.strptime(time_2, "%H:%M:%S")
                total_seconds = (time_2_struct - time_1_struct).total_seconds()
                min_sub = total_seconds // 60  # 计算打卡时长，单位为分钟

                wb["打卡"].cell(row=record_line, column=4, value=current_time)
                wb["打卡"].cell(row=record_line, column=1, value="ok")
                wb["打卡"].cell(row=record_line, column=7, value=min_sub)
            else:
                print("签到2")
                sendback = "ok"
                wa.append(
                    ["no", data_dict['I'], current_time, data_dict['W'], data_dict['ER'], dicline.get(data_dict['I']),
                 0])
            print("num : ", count_n)
        except :
            print("接收数据格式错误")
        try:
            wb.save(savepath)
            if (os.path.exists("D:/Web/files/operation/" + date_day + ".xlsx")):
                os.remove(os.path.join("D:/Web/files/operation/" + date_day + ".xlsx"))
            src = os.path.realpath(savepath)

            dst = os.path.realpath("D:/Web/files/operation/"+date_day+".xlsx")

            cmd = 'copy "%s" "%s"' % (src, dst)
            status = subprocess.check_output(['copy', src, dst], shell=True)
            # print("status: ", status.decode('utf-8'))
        except :
            sendback="no"
            print("请不要占用文件")
        # if(line_i==0):
        #     sheet.write(line_i + 1, 6, "0")
        #     book.save(savepath)
        # excel_file = xlrd.open_workbook(savepath)
        # wb = open_workbook(savepath)
        # read_sheet = wb.sheet_by_index(0)

        # line_i=len(open(savepath,encoding='gbk', errors="ignore").readlines())
        # if(read_sheet.cell(100,6).value!='NULL'):
        # print("read : ", read_sheet.cell(1, 6).value)
        # while True:
        #     if str(read_sheet.cell_value(line_i,6))!='a':
        #         line_i=1+line_i
        #     else :
        #         break
        #
        # print("line : ",line_i)
        # print("read : ",read_sheet.cell(line_i,6).value)

        print("date : ", current_time)
        # print("s: ",data_dict['S'])
        print("收到内容：", original_data)
        # print("发送地址：{}:{}".format(recv_data[1][0], recv_data[1][1]))
        # print("发送内容：", sendback.encode('GBK'))
        udp_socket.sendto(sendback.encode('GBK'), (recv_data[1][0], 8266))
        # udp_socket.sendto(data.encode('GBK'), (dest_ip, dest_port))


def main():
    udp_socket = socket(AF_INET, SOCK_DGRAM)
    local_addr = ('', 45555)
    udp_socket.bind(local_addr)
    while True:
        recv_msg(udp_socket)


if __name__ == '__main__':
    main()
